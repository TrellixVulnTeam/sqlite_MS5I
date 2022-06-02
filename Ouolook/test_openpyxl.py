
# ワークブックの読み込み
from operator import sub
from PySimpleGUI.PySimpleGUI import popup, popup_error
from openpyxl import load_workbook
import PySimpleGUI as sg
import win32com.client
import os 
import subprocess
import pyexcel as p
import pandas as pd
#ユーザーセッティング
settings = sg.UserSettings(filename="email_list",path=r"C:\Users\60837\Desktop\outlook")
settings.load()

#メール送信
def mail_send(To,CC,Bcc,Title,Body,):
    outlook = win32com.client.Dispatch("Outlook.Application")
    objMail = outlook.CreateItem(0) #MailItemオブジェクトのID
    
    #メール設定
    objMail.To = f"{To}" #宛先
    objMail.cc = f"{CC}" #CC
    objMail.Bcc = f"{Bcc}" #BCC
    objMail.Subject = f"{Title}" #Mailタイトル
    objMail.BodyFormat = 1 #(テキスト形式, リッチテキスト形式, HTML形式)
    objMail.Body = f"{Body}" #本文
    #objMail.Attachments.Add(r"{}".format(Add)) #添付ファイル
    objMail.Display(True) #MailItemオブジェクトを画面表示で確認する
    #objMail.Send() #メール即時送信

#大和化成の送り先選択肢
def selection_window():
    sg.theme("LightGray2")
    lay_selection=[
        [sg.Radio(text="内製",key="内製",group_id="A")],
        [sg.Radio(text="金型保全課",key="金型保全課",group_id="A")],
        [sg.Radio(text="矢野",key="矢野",group_id="A")],
        [sg.Button("OK",button_color=("white","blue"),key="OK")],
    ]
    window= sg.Window(title="送り先を選択",layout=lay_selection,size=(250,150))
    while True:
        event,values = window.read()
        if event == None:
            break
        if event == "キャンセル":
            break 
        if event == "OK":
            
            for select in values:
                if values[select] == True:
                    list_up = select
            window.close()
            return  list_up
            
                    

def main_prosess(file_path):

    #ワークブックを開く
    wb = load_workbook(r"{}".format(file_path))
    sheet_list = []
    for sheet in wb:
        sheet_list.append(sheet.title)#sheet_listにワークブックのシート一覧をappend
        

    sheet_value = {}
    for now in sheet_list:
        ws = wb[now]
        if ws["A11"].value =="金型修理・改善依頼書": #取得するシートを特定
            sheet_value["成形加工区"] = ws["M13"].value #成形メーカーを辞書に登録
            if ws["Q16"].value == "大和化成":#修理メーカーが大和化成の場合送り先を選択する
                name = selection_window()
                if name == None: #キャンセルを押したときの処理
                    mail_send(To=(settings[sheet_value["成形加工区"]]),CC="kenj-shibata@kojima-tns.co.jp; mai-ishikawa@kojima-tns.co.jp; yuusu-mori@kojima-tns.co.jp; ken-iwata@kojima-tns.co.jp; h-enomoto@kojima-tns.co.jp; katsuya-takahashi@kojima-tns.co.jp; y-ogino@kojima-tns.co.jp",Bcc="",Title="",Body="")
                else: #通常選択処理
                    mail_send(To=(settings[sheet_value["成形加工区"]],settings[name]),CC="kenj-shibata@kojima-tns.co.jp; mai-ishikawa@kojima-tns.co.jp; yuusu-mori@kojima-tns.co.jp; ken-iwata@kojima-tns.co.jp; h-enomoto@kojima-tns.co.jp; katsuya-takahashi@kojima-tns.co.jp; y-ogino@kojima-tns.co.jp",Bcc="",Title="",Body="")
                
            else:
                sheet_value["修理メーカー"] = ws["Q16"].value #修理メーカーを辞書に登録
                
                mail_send(To=(settings[sheet_value["成形加工区"]],settings[sheet_value["修理メーカー"]]),CC="kenj-shibata@kojima-tns.co.jp; mai-ishikawa@kojima-tns.co.jp; yuusu-mori@kojima-tns.co.jp; ken-iwata@kojima-tns.co.jp; h-enomoto@kojima-tns.co.jp; katsuya-takahashi@kojima-tns.co.jp; y-ogino@kojima-tns.co.jp",Bcc="",Title="",Body="")

#テーマ設定
sg.theme("Default")

#オプション設定
sg.set_options(use_ttk_buttons=True, dpi_awareness=True)

#フレームレイアウト
layoo = sg.Frame("その他",[
    [sg.Text("フォルダ内ファイル数"),sg.InputText(size=(10,1),key="file_count_out")],
    [sg.Text("拡張子'.xls'ファイル数"),sg.InputText(size=(10,1),key="xls_count_out",text_color="red")],
    [sg.Button("xlsファイルのみ表示", key= "display_xlsx")],
    [sg.Button("選択したファイルを削除",key="delete", button_color="red")],
    [sg.Button("メール作成",key="execution",pad=(80,20))],
],visible=False,key="soo")

#レイアウト
lay =[
    [sg.Text("指示書送付フォルダを選択"),sg.InputText(key="input_folder",),sg.FolderBrowse(button_text="フォルダを選択")],
    [sg.Button("検索",key="search"),sg.Button("ファイルを開く",key="file_open",visible=False,pad=(50,0),button_color="green"),sg.Button("フォルダを開く",key="folder_open",visible=False,pad=(50,0),button_color="green")],
    [sg.Listbox("",key="listbox",size=(50,10),visible=False),layoo],
      
]        

window = sg.Window("不具合指示書送付",lay)

while True:
    event,value = window.read()
    
    if event == None:
        break
    
    if event == "search": #検索ボタンを押したときの処理
        if value["input_folder"] == "":
            sg.popup("フォルダを選択してください")
            pass
        else:
            xls_list= []
            dir = os.listdir(value["input_folder"])
            for ii in dir:
                if ii.endswith(".xls") == True: #拡張子が.xlsのファイルはxls_listに移す
                    xls_list.append(ii)
            window["xls_count_out"].update(len(xls_list))
            window["file_count_out"].update(len(dir))
            window["listbox"].update(dir)
            window["listbox"].update(visible=True)
            window["file_open"].update(visible=True)
            window["folder_open"].update(visible=True)
            window["soo"].update(visible=True)
            
            
    if event == "folder_open": #フォルダを開くボタンを押したときの処理
        os.chdir(value["input_folder"])
        subprocess.Popen(["explorer", "."], shell=True)
        
        
    if event == "file_open": #ファイルを開くボタンを押したときの処理
        if window["listbox"].get() == []:
            
            sg.popup_error("選択してください")
            
        else:
            os.chdir(value["input_folder"])
            open_file_name = window["listbox"].get()
            
            subprocess.Popen(["C:\ProgramData\Microsoft\Windows\Start Menu\Programs\Microsoft Office 2013\Excel 2013.lnk",open_file_name[0]],shell=True)
    
    
    if event == "display_xlsx": #xlsファイルのみ表示ボタンを押したときの処理
        xls_list= []
        dir = os.listdir(value["input_folder"])
        for ii in dir:
            if ii.endswith(".xls") == True: #拡張子が.xlsのファイルはxls_listに移す
                xls_list.append(ii)
        window["listbox"].update(xls_list)
        
        
    if event == "delete": #選択したファイルを削除するボタンを押したときの処理
        if window["listbox"].get() == []:
            
            sg.popup_error("選択してください")
            
        else:
            os.chdir(value["input_folder"])
            open_file_name = window["listbox"].get()[0]
            sprit_text_file = os.path.splitext(open_file_name)[0]
            os.chdir(value["input_folder"])
            send = sg.popup_ok_cancel("ファイルを削除しますか？")
            if send == "OK":
                os.remove(open_file_name)
                dir = os.listdir(value["input_folder"])
                window["listbox"].update(dir)
                window["file_count_out"].update(len(dir))
                xls_list= []
                dir = os.listdir(value["input_folder"])
                for ii in dir:
                    if ii.endswith(".xls") == True: #拡張子が.xlsのファイルはxls_listに移す
                        xls_list.append(ii)
                window["xls_count_out"].update(len(xls_list))
            elif send == "Cancel":
                pass
            
            #df = pd.read_excel(r"{}".format(open_file_name))
            #df.to_excel(f"{sprit_text_file}.xlsx",encoding="shift-jis")
            
            #p.save_book_as(file_name="{}".format(open_file_name),
            #dest_file_name=f"{sprit_text_file}.xlsx")
        
    
    
    if event == "execution": #メール作成ボタンを押したときの処理
        if int(value["xls_count_out"]) >0: # .xlsファイルが含まれている場合はメール作成を中止する
            popup_error("拡張子が'.xls'のファイルが含まれています")
            pass
        else:   
            os.chdir(value["input_folder"]) #カレントディレクトリを作業フォルダに変更
            for i in dir:
                main_prosess(i)
            