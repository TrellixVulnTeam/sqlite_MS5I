import openpyxl
import PySimpleGUI as sg
def select(file_name):
    wb = openpyxl.load_workbook(file_name)   
    sheet_name = wb.sheetnames
    try:
        count = 0
        for i in  sheet_name:
            if i == "金型修理改善依頼書":
                ws =wb["金型修理改善依頼書"]
                count +=1

                
                

                name_list = {}

                #品番
                hinban = ws["C13"]
                #型番
                kataban = ws["C15"]
                #取数
                torisu = ws["F15"]
                #成型材質
                zaisitu = ws["I15"]
                #グレード
                grade = ws["I16"]
                #月産数
                gesan = ws["C17"]
                #製品形状
                keizyou = ws["I17"]
                #成型加工区
                kakouku = ws["M13"]
                #開始日
                kaishi = ws["M16"]
                #型納期
                nouki = ws["M17"]
                #製作メーカー
                seisaku = ws["Q15"]
                #修理メーカー
                syuuri = ws["Q16"]
                #修理区分
                kubun = ws["Q17"]
                #発行年月日
                hakou = ws["V13"]
                #手配ナンバー
                tehai_NO = ws["V15"]
                #ショットカウンタ
                shot_count = ws["Z22"]
                    
                name_list["品番"] = hinban.value 
                name_list["型番"] = kataban.value
                name_list["取数"] = torisu.value
                name_list["月産数"] = gesan.value
                name_list["成形加工区"] = kakouku.value
                name_list["開始日"] = str("{0:%Y/%m/%d}".format(kaishi.value))
                name_list["型納期"] = str("{0:%Y/%m/%d}".format(nouki.value))
                name_list["修理メーカー"] = syuuri.value
                name_list["手配ナンバー"] = tehai_NO.value
                name_list["ショットカウンタ"] = shot_count.value
        if count == 0:
            sg.popup_error("シート名が'金型修理改善依頼書'\nではありません")
            pass
                

        return name_list        
    except:
        pass




